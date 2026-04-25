"""
pipeline/exporter.py

Xuất kết quả OCR pipeline ra JSON và Excel.

Cấu trúc module:
  ── INPUT  : Định nghĩa kiểu dữ liệu đầu vào / đầu ra (TypedDict)
  ── LOGIC  : Các hàm xử lý nội bộ (group, build rows)
  ── OUTPUT : Ghi file (JSON, Excel)
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, TypedDict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


# =============================================================================
# INPUT — Định nghĩa cấu trúc dữ liệu
# =============================================================================
# Đây là contract dữ liệu nhận từ pipeline (table_parser.extract_structured_data)

class TableInput(TypedDict):
    """Một bảng được parse từ OCR text."""
    headers: List[str]              # tên các cột
    rows: List[Dict[str, str]]      # mỗi row là {header: value}


class OcrResult(TypedDict):
    """Kết quả đầy đủ của một ảnh/trang."""
    filename:    str                # tên file gốc
    ocr_text:    str                # văn bản OCR thô
    tables:      List[TableInput]   # danh sách bảng parse được
    text_lines:  List[str]          # dòng text ngoài bảng
    kv_pairs:    Dict[str, str]     # cặp key:value ngoài bảng
    table_count: int                # số bảng tìm thấy


# Kiểu dữ liệu nội bộ sau khi xử lý
class _SheetRow(TypedDict):
    """Một row trong sheet Excel, có thể là dòng text hoặc dòng dữ liệu."""
    _row_type: str       # "text" | "data"
    filename:  str
    _text:     str       # chỉ dùng cho _row_type="text"
    # + các key theo header bảng (chỉ cho _row_type="data")


class _SheetSpec(TypedDict):
    """Thông số đầy đủ của một sheet cần xuất."""
    full_headers: List[str]    # ["filename", "_text_content", col1, col2...]
    rows:         List[_SheetRow]


# =============================================================================
# LOGIC — Xử lý nội bộ (không ghi file, không đọc file)
# =============================================================================

# ── Hằng số sheet name ────────────────────────────────────────────────────────

_INVALID_CHARS = re.compile(r"[\\/*?\[\]:]")


def _sanitize_sheet_name(name: str) -> str:
    """Loại ký tự không hợp lệ, giới hạn 31 ký tự."""
    name = _INVALID_CHARS.sub("", name).strip()
    return name[:31] if name else "Sheet"


def _make_sheet_name(headers: List[str]) -> str:
    """Tạo tên sheet gợi nhớ từ 3 header đầu."""
    parts = [str(h)[:10] for h in headers[:3]]
    return _sanitize_sheet_name("_".join(parts)) or "Table"


def _unique_sheet_name(base: str, existing: Dict) -> str:
    """Tránh trùng tên bằng hậu tố _1, _2..."""
    if base not in existing:
        return base
    i = 1
    while f"{base[:28]}_{i}" in existing:
        i += 1
    return f"{base[:28]}_{i}"


# ── Bước 1: Phân nhóm kết quả theo cấu trúc bảng ────────────────────────────

def _group_by_structure(
    results: List[OcrResult],
) -> Tuple[Dict[str, _SheetSpec], List[Dict]]:
    """
    INPUT  : danh sách OcrResult từ pipeline
    OUTPUT : (sheets_spec, raw_rows)
               - sheets_spec: {sheet_name: _SheetSpec}  — chứa raw_data chờ Bước 2
               - raw_rows   : ảnh không có bảng nào     — fallback OCR_Raw sheet

    Logic  : gom các bảng có cùng cấu trúc header vào cùng 1 sheet
    """
    headers_to_sheet: Dict[Tuple, str] = {}
    sheets: Dict[str, Any] = {}
    raw_rows: List[Dict] = []

    for result in results:
        filename   = result.get("filename", "unknown")
        tables     = result.get("tables", [])
        text_lines = result.get("text_lines", [])
        ocr_text   = result.get("ocr_text", "")

        if not tables:
            raw_rows.append({"filename": filename, "ocr_text": ocr_text})
            continue

        for table in tables:
            headers = table.get("headers", [])
            if not headers:
                continue

            hkey = tuple(headers)
            if hkey not in headers_to_sheet:
                sname = _unique_sheet_name(_make_sheet_name(headers), sheets)
                headers_to_sheet[hkey] = sname
                sheets[sname] = {
                    "full_headers": ["filename", "_text_content"] + headers,
                    "raw_data": [],
                }

            sheets[headers_to_sheet[hkey]]["raw_data"].append({
                "filename":   filename,
                "text_lines": text_lines,
                "table":      table,
            })

    return sheets, raw_rows


# ── Bước 2: Dựng danh sách row phẳng cho từng sheet ─────────────────────────

def _build_rows(sheets: Dict[str, Any]) -> Dict[str, _SheetSpec]:
    """
    INPUT  : sheets với raw_data (từ Bước 1)
    OUTPUT : sheets với rows đã build (ready to write)

    Logic  : với mỗi entry trong raw_data:
               1. Text lines → _row_type="text"
               2. Table rows → _row_type="data"
    """
    for info in sheets.values():
        rows: List[_SheetRow] = []
        for entry in info["raw_data"]:
            filename   = entry["filename"]
            text_lines = entry["text_lines"]
            table      = entry["table"]

            for line in text_lines:
                rows.append({
                    "_row_type": "text",
                    "filename":  filename,
                    "_text":     line,
                })

            for row in table.get("rows", []):
                rows.append({
                    "_row_type":     "data",
                    "filename":      filename,
                    "_text_content": "",
                    **row,
                })

        info["rows"] = rows

    return sheets


# =============================================================================
# OUTPUT — Ghi file (JSON và Excel)
# =============================================================================

# ── Style constants ───────────────────────────────────────────────────────────

_STYLE = {
    "header_fill": PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
    "text_fill":   PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid"),
    "header_font": Font(bold=True, color="FFFFFF"),
    "text_font":   Font(italic=True, color="444444"),
    "center":      Alignment(horizontal="center", vertical="center"),
}


def _write_data_sheet(ws, full_headers: List[str], rows: List[_SheetRow]) -> None:
    """
    INPUT  : worksheet, full_headers, danh sách rows
    OUTPUT : worksheet được ghi đầy đủ (in-place)
    """
    # Header row
    ws.append(full_headers)
    for cell in ws[1]:
        cell.font      = _STYLE["header_font"]
        cell.fill      = _STYLE["header_fill"]
        cell.alignment = _STYLE["center"]

    # Data rows
    for row in rows:
        if row.get("_row_type") == "text":
            values = [row.get("filename", ""), row.get("_text", "")]
            values += [""] * (len(full_headers) - 2)
            ws.append(values)
            for cell in ws[ws.max_row]:
                cell.fill = _STYLE["text_fill"]
                cell.font = _STYLE["text_font"]
        else:
            ws.append([row.get(h, "") for h in full_headers])


def _write_raw_sheet(ws, raw_rows: List[Dict]) -> None:
    """
    INPUT  : worksheet, danh sách ảnh không có bảng
    OUTPUT : worksheet OCR_Raw được ghi (in-place)
    """
    ws.append(["filename", "ocr_text"])
    for cell in ws[1]:
        cell.font = _STYLE["header_font"]
        cell.fill = _STYLE["header_fill"]
    for row in raw_rows:
        ws.append([row["filename"], row["ocr_text"]])


# ── Public API ────────────────────────────────────────────────────────────────

def save_json(results: List[OcrResult], output_path: str) -> str:
    """
    INPUT  : danh sách OcrResult, đường dẫn file output
    OUTPUT : path đã lưu
    """
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    return output_path


def json_to_excel(results: List[OcrResult], output_path: str) -> str:
    """
    INPUT  : danh sách OcrResult (từ extract_structured_data)
    OUTPUT : đường dẫn file .xlsx đã tạo

    Luồng:
      Bước 1 [LOGIC] → _group_by_structure  — phân nhóm theo headers
      Bước 2 [LOGIC] → _build_rows          — dựng danh sách row phẳng
      Bước 3 [OUTPUT]→ _write_data_sheet     — ghi từng sheet vào workbook
      Bước 4 [OUTPUT]→ wb.save               — lưu file
    """
    # Bước 1: Phân nhóm
    sheets, raw_rows = _group_by_structure(results)

    # Bước 2: Dựng rows
    sheets = _build_rows(sheets)

    # Bước 3: Ghi workbook
    wb = Workbook()
    wb.remove(wb.active)

    for sname, info in sheets.items():
        ws = wb.create_sheet(title=sname[:31])
        _write_data_sheet(ws, info["full_headers"], info["rows"])

    if raw_rows:
        ws = wb.create_sheet(title="OCR_Raw")
        _write_raw_sheet(ws, raw_rows)

    if not wb.sheetnames:
        ws = wb.create_sheet("Empty")
        ws.append(["message"])
        ws.append(["No data extracted"])

    # Bước 4: Lưu
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
